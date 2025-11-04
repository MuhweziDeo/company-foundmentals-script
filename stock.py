#!/usr/bin/env python3
"""
Stock Fundamentals Analyzer
Collects company fundamental data using free APIs and displays it in the console.
Usage: python stock.py <TICKER>
"""

import sys
import argparse
from datetime import datetime
import yfinance as yf
import requests
from pathlib import Path

# ANSI color codes
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'
    WHITE = '\033[97m'


def get_financial_data(ticker):
    """Fetch financial data using yfinance"""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        
        # Get financial statements
        financials = stock.financials
        balance_sheet = stock.balance_sheet
        cashflow = stock.cashflow
        quarterly_financials = stock.quarterly_financials
        
        # Get historical data for price metrics
        hist = stock.history(period="max")  # Get all available historical data
        
        # Get earnings calendar for next earnings date
        next_earnings_date = None
        try:
            calendar = stock.calendar
            if calendar is not None and not calendar.empty:
                # Try to get next earnings date from calendar
                if 'Earnings Date' in calendar.index:
                    earnings_dates = calendar.loc['Earnings Date']
                    if len(earnings_dates) > 0:
                        # Get the first (most recent) earnings date
                        next_earnings_date = earnings_dates.iloc[0] if hasattr(earnings_dates, 'iloc') else earnings_dates[0]
                # Try alternative column names
                elif 'Earnings' in calendar.index:
                    earnings_dates = calendar.loc['Earnings']
                    if len(earnings_dates) > 0:
                        next_earnings_date = earnings_dates.iloc[0] if hasattr(earnings_dates, 'iloc') else earnings_dates[0]
        except Exception:
            pass
        
        # Fallback to info fields
        if not next_earnings_date:
            # Try various date fields that might contain earnings info
            for field in ['earningsDate', 'nextFiscalYearEnd', 'exDividendDate', 'mostRecentQuarter']:
                date_value = info.get(field)
                if date_value:
                    try:
                        # Try to parse if it's a timestamp
                        if isinstance(date_value, (int, float)):
                            from datetime import datetime as dt
                            next_earnings_date = dt.fromtimestamp(date_value)
                        else:
                            next_earnings_date = date_value
                        break
                    except:
                        continue
        
        return {
            'info': info,
            'financials': financials,
            'quarterly_financials': quarterly_financials,
            'balance_sheet': balance_sheet,
            'cashflow': cashflow,
            'stock': stock,
            'history': hist,
            'next_earnings_date': next_earnings_date
        }
    except Exception as e:
        print(f"{Colors.RED}Error fetching data: {e}{Colors.END}")
        return None


def calculate_rsi(prices, period=14):
    """Calculate Relative Strength Index (RSI)"""
    if prices is None or len(prices) < period + 1:
        return None
    
    try:
        import pandas as pd
        import numpy as np
        
        # Convert to pandas Series if not already
        if not isinstance(prices, pd.Series):
            prices = pd.Series(prices)
        
        # Calculate price changes
        delta = prices.diff()
        
        # Separate gains and losses
        gains = delta.where(delta > 0, 0)
        losses = -delta.where(delta < 0, 0)
        
        # Calculate average gains and losses using exponential moving average
        avg_gains = gains.ewm(span=period, adjust=False).mean()
        avg_losses = losses.ewm(span=period, adjust=False).mean()
        
        # Calculate RS and RSI
        rs = avg_gains / avg_losses
        rsi = 100 - (100 / (1 + rs))
        
        # Return the most recent RSI value
        return float(rsi.iloc[-1]) if not pd.isna(rsi.iloc[-1]) else None
    except Exception as e:
        return None


def get_rsi_signal(rsi):
    """Get RSI signal (oversold/overbought/neutral)"""
    if rsi is None:
        return None, "N/A"
    
    if rsi >= 70:
        return "OVERBOUGHT", Colors.RED
    elif rsi <= 30:
        return "OVERSOLD", Colors.GREEN
    else:
        return "NEUTRAL", Colors.YELLOW


def get_macro_data():
    """Fetch macroeconomic data"""
    try:
        # Using free API for interest rates (FRED API is free but requires registration)
        # For demo purposes, we'll use a simple approach
        # Note: In production, you'd want to use FRED API or similar
        return {}
    except Exception as e:
        return {}


def format_number(value):
    """Format large numbers in readable format"""
    if value is None:
        return "N/A"
    try:
        value = float(value)
        if abs(value) >= 1e12:
            return f"${value/1e12:.2f}T"
        elif abs(value) >= 1e9:
            return f"${value/1e9:.2f}B"
        elif abs(value) >= 1e6:
            return f"${value/1e6:.2f}M"
        elif abs(value) >= 1e3:
            return f"${value/1e3:.2f}K"
        else:
            return f"${value:.2f}"
    except:
        return str(value)


def format_percentage(value):
    """Format percentage values"""
    if value is None:
        return "N/A"
    try:
        return f"{float(value):.2f}%"
    except:
        return str(value)


def format_percentage_colored(value):
    """Format percentage values with color (green for positive, red for negative)"""
    if value is None:
        return f"{Colors.WHITE}N/A{Colors.END}"
    try:
        val = float(value)
        color = Colors.GREEN if val >= 0 else Colors.RED
        sign = "+" if val >= 0 else ""
        return f"{color}{sign}{val:.2f}%{Colors.END}"
    except:
        return f"{Colors.WHITE}{str(value)}{Colors.END}"


def format_ratio(value):
    """Format ratio values"""
    if value is None:
        return "N/A"
    try:
        return f"{float(value):.2f}"
    except:
        return str(value)


def format_price(value):
    """Format price values"""
    if value is None:
        return "N/A"
    try:
        return f"${float(value):.2f}"
    except:
        return str(value)


def display_price_data(data):
    """Display current price, all-time high/low, and YTD performance"""
    info = data['info']
    history = data['history']
    next_earnings_date = data.get('next_earnings_date')
    
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}0. PRICE INFORMATION{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    
    # Current price
    current_price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
    print(f"{Colors.BOLD}Current Price:{Colors.END} {Colors.YELLOW}{format_price(current_price)}{Colors.END}")
    
    # Calculate and display RSI
    rsi = None
    if not history.empty:
        try:
            # Get recent price data (last 30 days for RSI calculation)
            recent_prices = history['Close'].tail(30)
            if len(recent_prices) >= 15:  # Need at least 15 days for 14-period RSI
                rsi = calculate_rsi(recent_prices, period=14)
        except Exception:
            pass
    
    if rsi is not None:
        signal, signal_color = get_rsi_signal(rsi)
        rsi_color = Colors.RED if rsi >= 70 else Colors.GREEN if rsi <= 30 else Colors.YELLOW
        print(f"{Colors.BOLD}RSI (14-day):{Colors.END} {rsi_color}{format_ratio(rsi)}{Colors.END} {signal_color}({signal}){Colors.END}")
    else:
        print(f"{Colors.BOLD}RSI (14-day):{Colors.END} {Colors.WHITE}N/A{Colors.END}")
    
    # Next earnings date
    if next_earnings_date:
        try:
            # Format the date
            if hasattr(next_earnings_date, 'strftime'):
                earnings_date_str = next_earnings_date.strftime('%Y-%m-%d')
            elif isinstance(next_earnings_date, (list, tuple)) and len(next_earnings_date) > 0:
                earnings_date_str = str(next_earnings_date[0])
            else:
                earnings_date_str = str(next_earnings_date)
            
            # Calculate days until earnings
            try:
                if hasattr(next_earnings_date, 'strftime'):
                    earnings_date = next_earnings_date
                else:
                    from datetime import datetime as dt
                    earnings_date = dt.strptime(earnings_date_str, '%Y-%m-%d')
                
                days_until = (earnings_date.date() - datetime.now().date()).days if hasattr(earnings_date, 'date') else None
                days_str = f" ({days_until} days)" if days_until is not None and days_until >= 0 else ""
            except:
                days_str = ""
            
            print(f"{Colors.BOLD}Next Earnings Date:{Colors.END} {Colors.CYAN}{earnings_date_str}{days_str}{Colors.END}")
        except Exception:
            print(f"{Colors.BOLD}Next Earnings Date:{Colors.END} {Colors.WHITE}N/A{Colors.END}")
    else:
        print(f"{Colors.BOLD}Next Earnings Date:{Colors.END} {Colors.WHITE}N/A{Colors.END}")
    
    # Get YTD data
    current_year = datetime.now().year
    ytd_start = f"{current_year}-01-01"
    
    ytd_percent_change = None
    ytd_high = None
    ytd_low = None
    ytd_high_date = None
    ytd_low_date = None
    
    if not history.empty:
        try:
            # Get price at start of year
            ytd_prices = history[history.index >= ytd_start]
            if not ytd_prices.empty:
                year_start_price = ytd_prices['Close'].iloc[0]
                current_price_val = float(current_price) if current_price else None
                if current_price_val and year_start_price:
                    ytd_percent_change = ((current_price_val - year_start_price) / year_start_price) * 100
                
                # Get YTD high and low
                ytd_high = ytd_prices['High'].max()
                ytd_low = ytd_prices['Low'].min()
                
                # Get dates for YTD high and low
                ytd_high_date = ytd_prices[ytd_prices['High'] == ytd_high].index[0]
                ytd_low_date = ytd_prices[ytd_prices['Low'] == ytd_low].index[0]
        except Exception as e:
            pass
    
    # Alternative: try to get from info
    if ytd_percent_change is None:
        # Try to get from ytdReturn if available
        ytd_return = info.get('ytdReturn')
        if ytd_return:
            ytd_percent_change = ytd_return * 100
    
    print(f"{Colors.BOLD}YTD % Change:{Colors.END} {format_percentage_colored(ytd_percent_change)}")
    
    # YTD High and Low
    print(f"{Colors.BOLD}YTD High ({current_year}):{Colors.END} {Colors.GREEN}{format_price(ytd_high)}{Colors.END}", end="")
    if ytd_high_date is not None:
        print(f" {Colors.WHITE}(on {ytd_high_date.strftime('%Y-%m-%d')}){Colors.END}")
    else:
        print()
    
    print(f"{Colors.BOLD}YTD Low ({current_year}):{Colors.END} {Colors.RED}{format_price(ytd_low)}{Colors.END}", end="")
    if ytd_low_date is not None:
        print(f" {Colors.WHITE}(on {ytd_low_date.strftime('%Y-%m-%d')}){Colors.END}")
    else:
        print()
    
    # All-time high and low
    all_time_high = None
    all_time_low = None
    all_time_high_date = None
    all_time_low_date = None
    
    if not history.empty:
        try:
            # Get all-time high and low from historical data
            all_time_high = history['High'].max()
            all_time_low = history['Low'].min()
            
            # Get dates for all-time high and low
            all_time_high_date = history[history['High'] == all_time_high].index[0]
            all_time_low_date = history[history['Low'] == all_time_low].index[0]
        except Exception as e:
            pass
    
    # Fallback to 52-week high/low if available
    if all_time_high is None:
        all_time_high = info.get('fiftyTwoWeekHigh')
    if all_time_low is None:
        all_time_low = info.get('fiftyTwoWeekLow')
    
    print(f"{Colors.BOLD}All-Time High:{Colors.END} {Colors.GREEN}{format_price(all_time_high)}{Colors.END}", end="")
    if all_time_high_date is not None:
        print(f" {Colors.WHITE}(on {all_time_high_date.strftime('%Y-%m-%d')}){Colors.END}")
    else:
        print()
    
    print(f"{Colors.BOLD}All-Time Low:{Colors.END} {Colors.RED}{format_price(all_time_low)}{Colors.END}", end="")
    if all_time_low_date is not None:
        print(f" {Colors.WHITE}(on {all_time_low_date.strftime('%Y-%m-%d')}){Colors.END}")
    else:
        print()
    
    # Calculate distance from all-time high/low
    if current_price and all_time_high:
        try:
            current_price_val = float(current_price)
            ath_val = float(all_time_high)
            distance_from_ath = ((current_price_val - ath_val) / ath_val) * 100
            print(f"{Colors.BOLD}Distance from All-Time High:{Colors.END} {format_percentage_colored(distance_from_ath)}")
        except:
            pass
    
    if current_price and all_time_low:
        try:
            current_price_val = float(current_price)
            atl_val = float(all_time_low)
            distance_from_atl = ((current_price_val - atl_val) / atl_val) * 100
            print(f"{Colors.BOLD}Distance from All-Time Low:{Colors.END} {format_percentage_colored(distance_from_atl)}")
        except:
            pass


def display_financial_data(data):
    """Display financial metrics"""
    info = data['info']
    financials = data['financials']
    balance_sheet = data['balance_sheet']
    cashflow = data['cashflow']
    
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}1. FINANCIAL METRICS{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    
    # Revenue
    revenue = info.get('totalRevenue') or info.get('revenue')
    if revenue is None and not financials.empty:
        try:
            revenue = financials.loc['Total Revenue'].iloc[0] if 'Total Revenue' in financials.index else None
        except:
            pass
    print(f"{Colors.BOLD}Revenue (Sales):{Colors.END} {Colors.GREEN}{format_number(revenue)}{Colors.END}")
    
    # Earnings / Net Income
    net_income = info.get('netIncomeToCommon') or info.get('netIncome')
    if net_income is None and not financials.empty:
        try:
            net_income = financials.loc['Net Income'].iloc[0] if 'Net Income' in financials.index else None
        except:
            pass
    income_color = Colors.GREEN if (net_income and float(net_income) >= 0) else Colors.RED
    print(f"{Colors.BOLD}Earnings / Net Income:{Colors.END} {income_color}{format_number(net_income)}{Colors.END}")
    
    # Earnings Per Share (EPS)
    eps = info.get('trailingEps') or info.get('forwardEps')
    eps_color = Colors.GREEN if (eps and float(eps) >= 0) else Colors.RED
    print(f"{Colors.BOLD}Earnings Per Share (EPS):{Colors.END} {eps_color}{format_ratio(eps)}{Colors.END}")
    
    # Cash Flow
    operating_cashflow = info.get('operatingCashflow')
    if operating_cashflow is None and not cashflow.empty:
        try:
            operating_cashflow = cashflow.loc['Total Cash From Operating Activities'].iloc[0] if 'Total Cash From Operating Activities' in cashflow.index else None
        except:
            pass
    cashflow_color = Colors.GREEN if (operating_cashflow and float(operating_cashflow) >= 0) else Colors.RED
    print(f"{Colors.BOLD}Cash Flow (Operating):{Colors.END} {cashflow_color}{format_number(operating_cashflow)}{Colors.END}")
    
    # Debt-to-Equity Ratio
    debt_to_equity = info.get('debtToEquity')
    print(f"{Colors.BOLD}Debt-to-Equity Ratio:{Colors.END} {Colors.YELLOW}{format_ratio(debt_to_equity)}{Colors.END}")
    
    # Return on Equity (ROE)
    roe = info.get('returnOnEquity')
    if roe:
        roe = roe * 100
    roe_color = Colors.GREEN if (roe and float(roe) >= 0) else Colors.RED
    print(f"{Colors.BOLD}Return on Equity (ROE):{Colors.END} {roe_color}{format_percentage(roe)}{Colors.END}")
    
    # Gross Margin
    gross_margin = info.get('grossMargins')
    if gross_margin:
        gross_margin = gross_margin * 100
    print(f"{Colors.BOLD}Gross Margin:{Colors.END} {Colors.GREEN}{format_percentage(gross_margin)}{Colors.END}")
    
    # Net Margin
    net_margin = info.get('profitMargins')
    if net_margin:
        net_margin = net_margin * 100
    margin_color = Colors.GREEN if (net_margin and float(net_margin) >= 0) else Colors.RED
    print(f"{Colors.BOLD}Net Margin:{Colors.END} {margin_color}{format_percentage(net_margin)}{Colors.END}")


def display_valuation_data(data):
    """Display valuation metrics"""
    info = data['info']
    
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}2. VALUATION METRICS{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    
    # P/E Ratio
    pe_ratio = info.get('trailingPE') or info.get('forwardPE')
    print(f"{Colors.BOLD}P/E (Price-to-Earnings):{Colors.END} {Colors.YELLOW}{format_ratio(pe_ratio)}{Colors.END}")
    
    # P/B Ratio
    pb_ratio = info.get('priceToBook')
    print(f"{Colors.BOLD}P/B (Price-to-Book):{Colors.END} {Colors.YELLOW}{format_ratio(pb_ratio)}{Colors.END}")
    
    # P/S Ratio
    ps_ratio = info.get('priceToSalesTrailing12Months')
    print(f"{Colors.BOLD}P/S (Price-to-Sales):{Colors.END} {Colors.YELLOW}{format_ratio(ps_ratio)}{Colors.END}")
    
    # EV/EBITDA
    ev_ebitda = info.get('enterpriseToEbitda')
    print(f"{Colors.BOLD}EV/EBITDA:{Colors.END} {Colors.YELLOW}{format_ratio(ev_ebitda)}{Colors.END}")


def display_operation_data(data):
    """Display operational metrics"""
    info = data['info']
    
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}3. OPERATIONAL METRICS{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    
    # Market share - not directly available, show market cap as proxy
    market_cap = info.get('marketCap')
    print(f"{Colors.BOLD}Market Capitalization:{Colors.END} {Colors.GREEN}{format_number(market_cap)}{Colors.END}")
    print(f"{Colors.BOLD}Market Share:{Colors.END} {Colors.WHITE}N/A (requires industry-specific data){Colors.END}")
    
    # Business model - description
    business_summary = info.get('longBusinessSummary') or info.get('businessSummary')
    if business_summary:
        print(f"\n{Colors.BOLD}Business Model / Summary:{Colors.END}")
        # Truncate if too long
        summary = business_summary[:500] + "..." if len(business_summary) > 500 else business_summary
        print(f"  {Colors.WHITE}{summary}{Colors.END}")
    else:
        print(f"{Colors.BOLD}Business Model:{Colors.END} {Colors.WHITE}N/A{Colors.END}")
    
    # Competitive advantage - not directly available
    print(f"{Colors.BOLD}Competitive Advantage (Moat):{Colors.END} {Colors.WHITE}N/A (qualitative analysis required){Colors.END}")
    
    # Management quality - not directly available
    print(f"{Colors.BOLD}Management Quality:{Colors.END} {Colors.WHITE}N/A (qualitative analysis required){Colors.END}")
    
    # Product diversification - sector and industry
    sector = info.get('sector', 'N/A')
    industry = info.get('industry', 'N/A')
    print(f"{Colors.BOLD}Sector:{Colors.END} {Colors.CYAN}{sector}{Colors.END}")
    print(f"{Colors.BOLD}Industry:{Colors.END} {Colors.CYAN}{industry}{Colors.END}")
    
    # Number of employees as proxy for diversification
    employees = info.get('fullTimeEmployees')
    if employees:
        print(f"{Colors.BOLD}Full-Time Employees:{Colors.END} {Colors.YELLOW}{employees:,}{Colors.END}")


def get_earnings_trend(financials, quarterly_financials):
    """Extract earnings trend from financial statements"""
    earnings_trend = []
    
    # Try to get quarterly earnings first (more recent data)
    if quarterly_financials is not None and not quarterly_financials.empty:
        try:
            # Get Net Income from quarterly data
            if 'Net Income' in quarterly_financials.index:
                quarterly_data = quarterly_financials.loc['Net Income']
                for date, value in quarterly_data.items():
                    # Check for NaN (value != value checks for NaN)
                    if value is not None and not (isinstance(value, float) and value != value):
                        try:
                            earnings_trend.append({
                                'date': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date),
                                'earnings': float(value),
                                'period': 'Quarterly'
                            })
                        except (ValueError, TypeError):
                            pass
        except Exception:
            pass
    
    # Fallback to annual financials if quarterly not available
    if not earnings_trend and financials is not None and not financials.empty:
        try:
            if 'Net Income' in financials.index:
                annual_data = financials.loc['Net Income']
                for date, value in annual_data.items():
                    # Check for NaN (value != value checks for NaN)
                    if value is not None and not (isinstance(value, float) and value != value):
                        try:
                            earnings_trend.append({
                                'date': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date),
                                'earnings': float(value),
                                'period': 'Annual'
                            })
                        except (ValueError, TypeError):
                            pass
        except Exception:
            pass
    
    # Sort by date (most recent first)
    earnings_trend.sort(key=lambda x: x['date'], reverse=True)
    
    # Limit to last 8 periods for readability
    return earnings_trend[:8]


def display_earnings_trend(data):
    """Display earnings trend over time"""
    financials = data.get('financials')
    quarterly_financials = data.get('quarterly_financials')
    
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}5. EARNINGS TREND{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    
    earnings_trend = get_earnings_trend(financials, quarterly_financials)
    
    if not earnings_trend:
        print(f"{Colors.WHITE}Earnings trend data not available{Colors.END}")
        return
    
    print(f"\n{Colors.BOLD}Period{' ' * 20}Earnings{' ' * 15}Change{Colors.END}")
    print(f"{Colors.WHITE}{'-' * 80}{Colors.END}")
    
    for i, period in enumerate(earnings_trend):
        date_str = period['date']
        earnings = period['earnings']
        period_type = period['period']
        
        # Calculate change from previous period
        change_str = ""
        change_color = Colors.WHITE
        if i < len(earnings_trend) - 1:
            prev_earnings = earnings_trend[i + 1]['earnings']
            if prev_earnings != 0:
                change = ((earnings - prev_earnings) / abs(prev_earnings)) * 100
                change_color = Colors.GREEN if change >= 0 else Colors.RED
                sign = "+" if change >= 0 else ""
                change_str = f"{sign}{change:.2f}%"
            else:
                change_str = "N/A"
        
        earnings_color = Colors.GREEN if earnings >= 0 else Colors.RED
        earnings_formatted = format_number(earnings)
        
        # Format date nicely
        date_display = date_str[:10] if len(date_str) > 10 else date_str
        
        print(f"{Colors.WHITE}{date_display} ({period_type}){' ' * (20 - len(date_display) - len(period_type) - 2)}{earnings_color}{earnings_formatted}{' ' * (25 - len(earnings_formatted))}{change_color}{change_str}{Colors.END}")
    
    # Calculate overall trend
    if len(earnings_trend) >= 2:
        first_earnings = earnings_trend[-1]['earnings']
        latest_earnings = earnings_trend[0]['earnings']
        
        if first_earnings != 0:
            overall_change = ((latest_earnings - first_earnings) / abs(first_earnings)) * 100
            trend_color = Colors.GREEN if overall_change >= 0 else Colors.RED
            trend_direction = "↑ Increasing" if overall_change > 0 else "↓ Decreasing" if overall_change < 0 else "→ Stable"
            
            print(f"\n{Colors.BOLD}Overall Trend:{Colors.END} {trend_color}{trend_direction} ({overall_change:+.2f}%){Colors.END}")


def display_macro_sector_data(data):
    """Display macroeconomic and sector data"""
    info = data['info']
    
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}4. MACROECONOMIC AND SECTOR METRICS{Colors.END}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
    
    # Interest rates - not directly available from stock data
    print(f"{Colors.BOLD}Interest Rates:{Colors.END} {Colors.WHITE}N/A (requires external macro data API){Colors.END}")
    
    # Inflation - not directly available
    print(f"{Colors.BOLD}Inflation:{Colors.END} {Colors.WHITE}N/A (requires external macro data API){Colors.END}")
    
    # Consumer demand - use revenue growth as proxy
    revenue_growth = info.get('revenueGrowth')
    if revenue_growth:
        revenue_growth = revenue_growth * 100
    print(f"{Colors.BOLD}Revenue Growth (Demand Proxy):{Colors.END} {format_percentage_colored(revenue_growth)}")
    
    # Commodity prices - not directly available
    print(f"{Colors.BOLD}Commodity Prices:{Colors.END} {Colors.WHITE}N/A (requires commodity-specific data){Colors.END}")
    
    # Regulations / taxes - not directly available
    print(f"{Colors.BOLD}Regulations / Taxes:{Colors.END} {Colors.WHITE}N/A (requires regulatory research){Colors.END}")
    
    # Industry growth trends - use earnings growth as proxy
    earnings_growth = info.get('earningsGrowth')
    if earnings_growth:
        earnings_growth = earnings_growth * 100
    print(f"{Colors.BOLD}Earnings Growth (Industry Trend Proxy):{Colors.END} {format_percentage_colored(earnings_growth)}")
    
    # Additional sector info
    beta = info.get('beta')
    print(f"{Colors.BOLD}Beta (Market Correlation):{Colors.END} {Colors.YELLOW}{format_ratio(beta)}{Colors.END}")


def collect_all_data(data, ticker):
    """Collect all data in a structured format for export"""
    info = data['info']
    financials = data['financials']
    quarterly_financials = data.get('quarterly_financials')
    balance_sheet = data['balance_sheet']
    cashflow = data['cashflow']
    history = data['history']
    next_earnings_date = data.get('next_earnings_date')
    
    # Get earnings trend
    earnings_trend = get_earnings_trend(financials, quarterly_financials)
    
    # Calculate RSI
    rsi = None
    if not history.empty:
        try:
            recent_prices = history['Close'].tail(30)
            if len(recent_prices) >= 15:
                rsi = calculate_rsi(recent_prices, period=14)
        except Exception:
            pass
    
    # Format next earnings date
    earnings_date_str = None
    days_until_earnings = None
    if next_earnings_date:
        try:
            if hasattr(next_earnings_date, 'strftime'):
                earnings_date_str = next_earnings_date.strftime('%Y-%m-%d')
                days_until_earnings = (next_earnings_date.date() - datetime.now().date()).days if hasattr(next_earnings_date, 'date') else None
            else:
                earnings_date_str = str(next_earnings_date)
        except Exception:
            pass
    
    # Get YTD data
    current_year = datetime.now().year
    ytd_start = f"{current_year}-01-01"
    
    ytd_percent_change = None
    ytd_high = None
    ytd_low = None
    ytd_high_date = None
    ytd_low_date = None
    
    if not history.empty:
        try:
            ytd_prices = history[history.index >= ytd_start]
            if not ytd_prices.empty:
                year_start_price = ytd_prices['Close'].iloc[0]
                current_price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
                current_price_val = float(current_price) if current_price else None
                if current_price_val and year_start_price:
                    ytd_percent_change = ((current_price_val - year_start_price) / year_start_price) * 100
                
                ytd_high = ytd_prices['High'].max()
                ytd_low = ytd_prices['Low'].min()
                ytd_high_date = ytd_prices[ytd_prices['High'] == ytd_high].index[0]
                ytd_low_date = ytd_prices[ytd_prices['Low'] == ytd_low].index[0]
        except:
            pass
    
    # All-time high/low
    all_time_high = None
    all_time_low = None
    all_time_high_date = None
    all_time_low_date = None
    
    if not history.empty:
        try:
            all_time_high = history['High'].max()
            all_time_low = history['Low'].min()
            all_time_high_date = history[history['High'] == all_time_high].index[0]
            all_time_low_date = history[history['Low'] == all_time_low].index[0]
        except:
            pass
    
    if all_time_high is None:
        all_time_high = info.get('fiftyTwoWeekHigh')
    if all_time_low is None:
        all_time_low = info.get('fiftyTwoWeekLow')
    
    # Revenue
    revenue = info.get('totalRevenue') or info.get('revenue')
    if revenue is None and not financials.empty:
        try:
            revenue = financials.loc['Total Revenue'].iloc[0] if 'Total Revenue' in financials.index else None
        except:
            pass
    
    # Net Income
    net_income = info.get('netIncomeToCommon') or info.get('netIncome')
    if net_income is None and not financials.empty:
        try:
            net_income = financials.loc['Net Income'].iloc[0] if 'Net Income' in financials.index else None
        except:
            pass
    
    # Operating Cash Flow
    operating_cashflow = info.get('operatingCashflow')
    if operating_cashflow is None and not cashflow.empty:
        try:
            operating_cashflow = cashflow.loc['Total Cash From Operating Activities'].iloc[0] if 'Total Cash From Operating Activities' in cashflow.index else None
        except:
            pass
    
    # Revenue and Earnings Growth
    revenue_growth = info.get('revenueGrowth')
    if revenue_growth:
        revenue_growth = revenue_growth * 100
    
    earnings_growth = info.get('earningsGrowth')
    if earnings_growth:
        earnings_growth = earnings_growth * 100
    
    # ROE and Margins
    roe = info.get('returnOnEquity')
    if roe:
        roe = roe * 100
    
    gross_margin = info.get('grossMargins')
    if gross_margin:
        gross_margin = gross_margin * 100
    
    net_margin = info.get('profitMargins')
    if net_margin:
        net_margin = net_margin * 100
    
    return {
        'ticker': ticker,
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'price_info': {
            'current_price': info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose'),
            'rsi': rsi,
            'rsi_signal': get_rsi_signal(rsi)[0] if rsi is not None else None,
            'next_earnings_date': earnings_date_str,
            'days_until_earnings': days_until_earnings,
            'ytd_percent_change': ytd_percent_change,
            'ytd_high': ytd_high,
            'ytd_high_date': ytd_high_date.strftime('%Y-%m-%d') if ytd_high_date else None,
            'ytd_low': ytd_low,
            'ytd_low_date': ytd_low_date.strftime('%Y-%m-%d') if ytd_low_date else None,
            'all_time_high': all_time_high,
            'all_time_high_date': all_time_high_date.strftime('%Y-%m-%d') if all_time_high_date else None,
            'all_time_low': all_time_low,
            'all_time_low_date': all_time_low_date.strftime('%Y-%m-%d') if all_time_low_date else None,
        },
        'financial': {
            'revenue': revenue,
            'net_income': net_income,
            'eps': info.get('trailingEps') or info.get('forwardEps'),
            'cash_flow': operating_cashflow,
            'debt_to_equity': info.get('debtToEquity'),
            'roe': roe,
            'gross_margin': gross_margin,
            'net_margin': net_margin,
        },
        'valuation': {
            'pe_ratio': info.get('trailingPE') or info.get('forwardPE'),
            'pb_ratio': info.get('priceToBook'),
            'ps_ratio': info.get('priceToSalesTrailing12Months'),
            'ev_ebitda': info.get('enterpriseToEbitda'),
        },
        'operational': {
            'market_cap': info.get('marketCap'),
            'sector': info.get('sector', 'N/A'),
            'industry': info.get('industry', 'N/A'),
            'employees': info.get('fullTimeEmployees'),
            'business_summary': info.get('longBusinessSummary') or info.get('businessSummary'),
        },
        'macro_sector': {
            'revenue_growth': revenue_growth,
            'earnings_growth': earnings_growth,
            'beta': info.get('beta'),
        },
        'earnings_trend': earnings_trend
    }


def export_to_excel(data_dict, output_path):
    """Export data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Analysis"
        
        # Header
        ws['A1'] = f"STOCK FUNDAMENTALS ANALYSIS FOR: {data_dict['ticker']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Analysis Date: {data_dict['analysis_date']}"
        
        row = 4
        
        # Price Information
        ws[f'A{row}'] = "PRICE INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        price_info = data_dict['price_info']
        rsi_value = price_info.get('rsi')
        rsi_signal = price_info.get('rsi_signal', 'N/A')
        rsi_display = f"{rsi_value:.2f} ({rsi_signal})" if rsi_value is not None else 'N/A'
        
        next_earnings = price_info.get('next_earnings_date')
        days_until = price_info.get('days_until_earnings')
        earnings_display = next_earnings if next_earnings else 'N/A'
        if days_until is not None and days_until >= 0:
            earnings_display += f" ({days_until} days)"
        
        price_data = [
            ['Current Price', f"${price_info['current_price']:.2f}" if price_info['current_price'] else 'N/A'],
            ['RSI (14-day)', rsi_display],
            ['Next Earnings Date', earnings_display],
            ['YTD % Change', f"{price_info['ytd_percent_change']:.2f}%" if price_info['ytd_percent_change'] is not None else 'N/A'],
            ['YTD High', f"${price_info['ytd_high']:.2f}" if price_info['ytd_high'] else 'N/A', price_info['ytd_high_date']],
            ['YTD Low', f"${price_info['ytd_low']:.2f}" if price_info['ytd_low'] else 'N/A', price_info['ytd_low_date']],
            ['All-Time High', f"${price_info['all_time_high']:.2f}" if price_info['all_time_high'] else 'N/A', price_info['all_time_high_date']],
            ['All-Time Low', f"${price_info['all_time_low']:.2f}" if price_info['all_time_low'] else 'N/A', price_info['all_time_low_date']],
        ]
        
        for item in price_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if len(item) > 1 else ''
            if len(item) > 2 and item[2]:
                ws[f'C{row}'] = f"({item[2]})"
            row += 1
        
        row += 1
        
        # Financial Metrics
        ws[f'A{row}'] = "FINANCIAL METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        financial = data_dict['financial']
        financial_data = [
            ['Revenue (Sales)', format_number(financial['revenue'])],
            ['Earnings / Net Income', format_number(financial['net_income'])],
            ['Earnings Per Share (EPS)', format_ratio(financial['eps'])],
            ['Cash Flow (Operating)', format_number(financial['cash_flow'])],
            ['Debt-to-Equity Ratio', format_ratio(financial['debt_to_equity'])],
            ['Return on Equity (ROE)', f"{financial['roe']:.2f}%" if financial['roe'] is not None else 'N/A'],
            ['Gross Margin', f"{financial['gross_margin']:.2f}%" if financial['gross_margin'] is not None else 'N/A'],
            ['Net Margin', f"{financial['net_margin']:.2f}%" if financial['net_margin'] is not None else 'N/A'],
        ]
        
        for item in financial_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        row += 1
        
        # Valuation Metrics
        ws[f'A{row}'] = "VALUATION METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        valuation = data_dict['valuation']
        valuation_data = [
            ['P/E (Price-to-Earnings)', format_ratio(valuation['pe_ratio'])],
            ['P/B (Price-to-Book)', format_ratio(valuation['pb_ratio'])],
            ['P/S (Price-to-Sales)', format_ratio(valuation['ps_ratio'])],
            ['EV/EBITDA', format_ratio(valuation['ev_ebitda'])],
        ]
        
        for item in valuation_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        row += 1
        
        # Operational Metrics
        ws[f'A{row}'] = "OPERATIONAL METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        operational = data_dict['operational']
        operational_data = [
            ['Market Capitalization', format_number(operational['market_cap'])],
            ['Sector', operational['sector']],
            ['Industry', operational['industry']],
            ['Full-Time Employees', f"{operational['employees']:,}" if operational['employees'] else 'N/A'],
        ]
        
        for item in operational_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        if operational['business_summary']:
            row += 1
            ws[f'A{row}'] = "Business Summary"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            summary = operational['business_summary'][:500] + "..." if len(operational['business_summary']) > 500 else operational['business_summary']
            ws[f'A{row}'] = summary
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        row += 1
        
        # Macro/Sector Metrics
        ws[f'A{row}'] = "MACROECONOMIC AND SECTOR METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        macro = data_dict['macro_sector']
        macro_data = [
            ['Revenue Growth', f"{macro['revenue_growth']:.2f}%" if macro['revenue_growth'] is not None else 'N/A'],
            ['Earnings Growth', f"{macro['earnings_growth']:.2f}%" if macro['earnings_growth'] is not None else 'N/A'],
            ['Beta (Market Correlation)', format_ratio(macro['beta'])],
        ]
        
        for item in macro_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        row += 1
        
        # Earnings Trend
        earnings_trend = data_dict.get('earnings_trend', [])
        if earnings_trend:
            ws[f'A{row}'] = "EARNINGS TREND"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
            row += 1
            
            ws[f'A{row}'] = "Period"
            ws[f'B{row}'] = "Earnings"
            ws[f'C{row}'] = "Change"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
            row += 1
            
            for i, period in enumerate(earnings_trend):
                date_str = period['date']
                earnings = period['earnings']
                period_type = period['period']
                
                ws[f'A{row}'] = f"{date_str} ({period_type})"
                ws[f'B{row}'] = format_number(earnings)
                
                # Calculate change
                if i < len(earnings_trend) - 1:
                    prev_earnings = earnings_trend[i + 1]['earnings']
                    if prev_earnings != 0:
                        change = ((earnings - prev_earnings) / abs(prev_earnings)) * 100
                        ws[f'C{row}'] = f"{change:+.2f}%"
                    else:
                        ws[f'C{row}'] = "N/A"
                else:
                    ws[f'C{row}'] = "-"
                
                row += 1
            
            # Overall trend
            if len(earnings_trend) >= 2:
                first_earnings = earnings_trend[-1]['earnings']
                latest_earnings = earnings_trend[0]['earnings']
                if first_earnings != 0:
                    overall_change = ((latest_earnings - first_earnings) / abs(first_earnings)) * 100
                    row += 1
                    ws[f'A{row}'] = "Overall Trend"
                    ws[f'A{row}'].font = Font(bold=True)
                    trend_direction = "↑ Increasing" if overall_change > 0 else "↓ Decreasing" if overall_change < 0 else "→ Stable"
                    ws[f'B{row}'] = f"{trend_direction} ({overall_change:+.2f}%)"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        
        wb.save(output_path)
        return True
    except ImportError:
        print(f"{Colors.RED}Error: openpyxl and pandas are required for Excel export. Install with: pip install openpyxl pandas{Colors.END}")
        return False
    except Exception as e:
        print(f"{Colors.RED}Error exporting to Excel: {e}{Colors.END}")
        return False


def export_to_pdf(data_dict, output_path):
    """Export data to PDF file"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0066CC'),
            spaceAfter=30,
        )
        
        title = Paragraph(f"STOCK FUNDAMENTALS ANALYSIS FOR: {data_dict['ticker']}", title_style)
        story.append(title)
        story.append(Paragraph(f"Analysis Date: {data_dict['analysis_date']}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Helper function to create section
        def add_section(title_text, data_list):
            story.append(Paragraph(f"<b>{title_text}</b>", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            table_data = [['Metric', 'Value']]
            for item in data_list:
                if isinstance(item, list) and len(item) >= 2:
                    metric = item[0]
                    value = item[1] if item[1] is not None else 'N/A'
                    if len(item) > 2:
                        value = f"{value} {item[2]}"
                    table_data.append([metric, str(value)])
            
            table = Table(table_data, colWidths=[3*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCFFFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
        
        # Price Information
        price_info = data_dict['price_info']
        rsi_value = price_info.get('rsi')
        rsi_signal = price_info.get('rsi_signal', 'N/A')
        rsi_display = f"{rsi_value:.2f} ({rsi_signal})" if rsi_value is not None else 'N/A'
        
        next_earnings = price_info.get('next_earnings_date')
        days_until = price_info.get('days_until_earnings')
        earnings_display = next_earnings if next_earnings else 'N/A'
        if days_until is not None and days_until >= 0:
            earnings_display += f" ({days_until} days)"
        
        price_data = [
            ['Current Price', f"${price_info['current_price']:.2f}" if price_info['current_price'] else 'N/A'],
            ['RSI (14-day)', rsi_display],
            ['Next Earnings Date', earnings_display],
            ['YTD % Change', f"{price_info['ytd_percent_change']:.2f}%" if price_info['ytd_percent_change'] else 'N/A'],
            ['YTD High', f"${price_info['ytd_high']:.2f}" if price_info['ytd_high'] else 'N/A', f"(on {price_info['ytd_high_date']})" if price_info['ytd_high_date'] else ''],
            ['YTD Low', f"${price_info['ytd_low']:.2f}" if price_info['ytd_low'] else 'N/A', f"(on {price_info['ytd_low_date']})" if price_info['ytd_low_date'] else ''],
            ['All-Time High', f"${price_info['all_time_high']:.2f}" if price_info['all_time_high'] else 'N/A', f"(on {price_info['all_time_high_date']})" if price_info['all_time_high_date'] else ''],
            ['All-Time Low', f"${price_info['all_time_low']:.2f}" if price_info['all_time_low'] else 'N/A', f"(on {price_info['all_time_low_date']})" if price_info['all_time_low_date'] else ''],
        ]
        add_section("PRICE INFORMATION", price_data)
        
        # Financial Metrics
        financial = data_dict['financial']
        financial_data = [
            ['Revenue (Sales)', format_number(financial['revenue'])],
            ['Earnings / Net Income', format_number(financial['net_income'])],
            ['Earnings Per Share (EPS)', format_ratio(financial['eps'])],
            ['Cash Flow (Operating)', format_number(financial['cash_flow'])],
            ['Debt-to-Equity Ratio', format_ratio(financial['debt_to_equity'])],
            ['Return on Equity (ROE)', f"{financial['roe']:.2f}%" if financial['roe'] else 'N/A'],
            ['Gross Margin', f"{financial['gross_margin']:.2f}%" if financial['gross_margin'] else 'N/A'],
            ['Net Margin', f"{financial['net_margin']:.2f}%" if financial['net_margin'] else 'N/A'],
        ]
        add_section("FINANCIAL METRICS", financial_data)
        
        # Valuation Metrics
        valuation = data_dict['valuation']
        valuation_data = [
            ['P/E (Price-to-Earnings)', format_ratio(valuation['pe_ratio'])],
            ['P/B (Price-to-Book)', format_ratio(valuation['pb_ratio'])],
            ['P/S (Price-to-Sales)', format_ratio(valuation['ps_ratio'])],
            ['EV/EBITDA', format_ratio(valuation['ev_ebitda'])],
        ]
        add_section("VALUATION METRICS", valuation_data)
        
        # Operational Metrics
        operational = data_dict['operational']
        operational_data = [
            ['Market Capitalization', format_number(operational['market_cap'])],
            ['Sector', operational['sector']],
            ['Industry', operational['industry']],
            ['Full-Time Employees', f"{operational['employees']:,}" if operational['employees'] else 'N/A'],
        ]
        add_section("OPERATIONAL METRICS", operational_data)
        
        if operational['business_summary']:
            story.append(Paragraph("<b>Business Summary</b>", styles['Heading3']))
            summary = operational['business_summary'][:1000] + "..." if len(operational['business_summary']) > 1000 else operational['business_summary']
            story.append(Paragraph(summary, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Macro/Sector Metrics
        macro = data_dict['macro_sector']
        macro_data = [
            ['Revenue Growth', f"{macro['revenue_growth']:.2f}%" if macro['revenue_growth'] else 'N/A'],
            ['Earnings Growth', f"{macro['earnings_growth']:.2f}%" if macro['earnings_growth'] else 'N/A'],
            ['Beta (Market Correlation)', format_ratio(macro['beta'])],
        ]
        add_section("MACROECONOMIC AND SECTOR METRICS", macro_data)
        
        # Earnings Trend
        earnings_trend = data_dict.get('earnings_trend', [])
        if earnings_trend:
            story.append(Paragraph("<b>EARNINGS TREND</b>", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            trend_table_data = [['Period', 'Earnings', 'Change']]
            for i, period in enumerate(earnings_trend):
                date_str = period['date']
                earnings = period['earnings']
                period_type = period['period']
                
                change_str = ""
                if i < len(earnings_trend) - 1:
                    prev_earnings = earnings_trend[i + 1]['earnings']
                    if prev_earnings != 0:
                        change = ((earnings - prev_earnings) / abs(prev_earnings)) * 100
                        change_str = f"{change:+.2f}%"
                    else:
                        change_str = "N/A"
                else:
                    change_str = "-"
                
                trend_table_data.append([
                    f"{date_str} ({period_type})",
                    format_number(earnings),
                    change_str
                ])
            
            trend_table = Table(trend_table_data, colWidths=[2.5*inch, 2.5*inch, 1*inch])
            trend_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCFFFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(trend_table)
            
            # Overall trend
            if len(earnings_trend) >= 2:
                first_earnings = earnings_trend[-1]['earnings']
                latest_earnings = earnings_trend[0]['earnings']
                if first_earnings != 0:
                    overall_change = ((latest_earnings - first_earnings) / abs(first_earnings)) * 100
                    trend_direction = "↑ Increasing" if overall_change > 0 else "↓ Decreasing" if overall_change < 0 else "→ Stable"
                    story.append(Spacer(1, 0.1*inch))
                    story.append(Paragraph(f"<b>Overall Trend:</b> {trend_direction} ({overall_change:+.2f}%)", styles['Normal']))
            
            story.append(Spacer(1, 0.2*inch))
        
        doc.build(story)
        return True
    except ImportError:
        print(f"{Colors.RED}Error: reportlab is required for PDF export. Install with: pip install reportlab{Colors.END}")
        return False
    except Exception as e:
        print(f"{Colors.RED}Error exporting to PDF: {e}{Colors.END}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Stock Fundamentals Analyzer - Collects company fundamental data'
    )
    parser.add_argument('ticker', type=str.upper, help='Stock ticker symbol (e.g., AAPL, MSFT)')
    parser.add_argument('-o', '--output', type=str, help='Output file path (supports .xlsx or .pdf extension)')
    
    args = parser.parse_args()
    ticker = args.ticker
    output_path = args.output
    
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}STOCK FUNDAMENTALS ANALYSIS FOR: {Colors.YELLOW}{ticker}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*80}{Colors.END}")
    print(f"{Colors.WHITE}Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}{Colors.END}")
    
    # Fetch data
    print(f"\n{Colors.YELLOW}Fetching data for {ticker}...{Colors.END}")
    data = get_financial_data(ticker)
    
    if data is None:
        print(f"\n{Colors.RED}{Colors.BOLD}Error: Could not fetch data for ticker {ticker}{Colors.END}")
        print(f"{Colors.RED}Please verify the ticker symbol is correct.{Colors.END}")
        sys.exit(1)
    
    # Display all sections
    display_price_data(data)
    display_financial_data(data)
    display_valuation_data(data)
    display_operation_data(data)
    display_macro_sector_data(data)
    display_earnings_trend(data)
    
    # Export to file if requested
    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Collect all data for export
        export_data = collect_all_data(data, ticker)
        
        # Determine format from extension
        if output_path.suffix.lower() == '.xlsx':
            print(f"\n{Colors.YELLOW}📊 Exporting to Excel: {output_path}{Colors.END}")
            if export_to_excel(export_data, str(output_path)):
                print(f"{Colors.GREEN}✓ Successfully exported to {output_path}{Colors.END}")
            else:
                sys.exit(1)
        elif output_path.suffix.lower() == '.pdf':
            print(f"\n{Colors.YELLOW}📄 Exporting to PDF: {output_path}{Colors.END}")
            if export_to_pdf(export_data, str(output_path)):
                print(f"{Colors.GREEN}✓ Successfully exported to {output_path}{Colors.END}")
            else:
                sys.exit(1)
        else:
            print(f"{Colors.RED}Error: Unsupported file format. Use .xlsx or .pdf extension.{Colors.END}")
            sys.exit(1)
    
    print(f"\n{Colors.BOLD}{Colors.GREEN}{'='*80}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.GREEN}Analysis Complete{Colors.END}")
    print(f"{Colors.BOLD}{Colors.GREEN}{'='*80}{Colors.END}\n")
    
    # Note about data availability
    print(f"{Colors.YELLOW}Note: Some metrics (marked as N/A) require:{Colors.END}")
    print(f"{Colors.WHITE}  - Qualitative analysis (business model, competitive advantage, management){Colors.END}")
    print(f"{Colors.WHITE}  - External macroeconomic data APIs (interest rates, inflation){Colors.END}")
    print(f"{Colors.WHITE}  - Industry-specific research (market share, regulations){Colors.END}")
    print()


if __name__ == "__main__":
    main()

