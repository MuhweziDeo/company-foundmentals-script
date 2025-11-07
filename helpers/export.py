"""
Export functions for data export to Excel and PDF
"""

from datetime import datetime
from .constants import Colors
from .formatters import format_number, format_ratio, format_price
from .calculators import (
    calculate_rsi,
    get_rsi_signal,
    get_earnings_trend,
    get_analyst_ratings
)

def collect_all_data(data, ticker):
    """Collect all data in a structured format for export"""
    info = data['info']
    financials = data['financials']
    quarterly_financials = data.get('quarterly_financials')
    balance_sheet = data['balance_sheet']
    cashflow = data['cashflow']
    history = data['history']
    next_earnings_date = data.get('next_earnings_date')
    
    # Get earnings trend
    earnings_trend = get_earnings_trend(financials, quarterly_financials)
    
    # Get analyst ratings
    recommendations = data.get('recommendations')
    current_price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
    analyst_data = get_analyst_ratings(info, recommendations, current_price)
    
    # Calculate RSI
    rsi = None
    if not history.empty:
        try:
            recent_prices = history['Close'].tail(30)
            if len(recent_prices) >= 15:
                rsi = calculate_rsi(recent_prices, period=14)
        except Exception:
            pass
    
    # Format next earnings date
    earnings_date_str = None
    days_until_earnings = None
    if next_earnings_date:
        try:
            if hasattr(next_earnings_date, 'strftime'):
                earnings_date_str = next_earnings_date.strftime('%Y-%m-%d')
                days_until_earnings = (next_earnings_date.date() - datetime.now().date()).days if hasattr(next_earnings_date, 'date') else None
            else:
                earnings_date_str = str(next_earnings_date)
        except Exception:
            pass
    
    # Get YTD data
    current_year = datetime.now().year
    ytd_start = f"{current_year}-01-01"
    
    ytd_percent_change = None
    ytd_high = None
    ytd_low = None
    ytd_high_date = None
    ytd_low_date = None
    
    if not history.empty:
        try:
            ytd_prices = history[history.index >= ytd_start]
            if not ytd_prices.empty:
                year_start_price = ytd_prices['Close'].iloc[0]
                current_price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
                current_price_val = float(current_price) if current_price else None
                if current_price_val and year_start_price:
                    ytd_percent_change = ((current_price_val - year_start_price) / year_start_price) * 100
                
                ytd_high = ytd_prices['High'].max()
                ytd_low = ytd_prices['Low'].min()
                ytd_high_date = ytd_prices[ytd_prices['High'] == ytd_high].index[0]
                ytd_low_date = ytd_prices[ytd_prices['Low'] == ytd_low].index[0]
        except:
            pass
    
    # All-time high/low
    all_time_high = None
    all_time_low = None
    all_time_high_date = None
    all_time_low_date = None
    
    if not history.empty:
        try:
            all_time_high = history['High'].max()
            all_time_low = history['Low'].min()
            all_time_high_date = history[history['High'] == all_time_high].index[0]
            all_time_low_date = history[history['Low'] == all_time_low].index[0]
        except:
            pass
    
    if all_time_high is None:
        all_time_high = info.get('fiftyTwoWeekHigh')
    if all_time_low is None:
        all_time_low = info.get('fiftyTwoWeekLow')
    
    # Revenue
    revenue = info.get('totalRevenue') or info.get('revenue')
    if revenue is None and not financials.empty:
        try:
            revenue = financials.loc['Total Revenue'].iloc[0] if 'Total Revenue' in financials.index else None
        except:
            pass
    
    # Net Income
    net_income = info.get('netIncomeToCommon') or info.get('netIncome')
    if net_income is None and not financials.empty:
        try:
            net_income = financials.loc['Net Income'].iloc[0] if 'Net Income' in financials.index else None
        except:
            pass
    
    # Operating Cash Flow
    operating_cashflow = info.get('operatingCashflow')
    if operating_cashflow is None and not cashflow.empty:
        try:
            operating_cashflow = cashflow.loc['Total Cash From Operating Activities'].iloc[0] if 'Total Cash From Operating Activities' in cashflow.index else None
        except:
            pass
    
    # Revenue and Earnings Growth
    revenue_growth = info.get('revenueGrowth')
    if revenue_growth:
        revenue_growth = revenue_growth * 100
    
    earnings_growth = info.get('earningsGrowth')
    if earnings_growth:
        earnings_growth = earnings_growth * 100
    
    # ROE and Margins
    roe = info.get('returnOnEquity')
    if roe:
        roe = roe * 100
    
    gross_margin = info.get('grossMargins')
    if gross_margin:
        gross_margin = gross_margin * 100
    
    net_margin = info.get('profitMargins')
    if net_margin:
        net_margin = net_margin * 100
    
    return {
        'ticker': ticker,
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'price_info': {
            'current_price': info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose'),
            'rsi': rsi,
            'rsi_signal': get_rsi_signal(rsi)[0] if rsi is not None else None,
            'next_earnings_date': earnings_date_str,
            'days_until_earnings': days_until_earnings,
            'ytd_percent_change': ytd_percent_change,
            'ytd_high': ytd_high,
            'ytd_high_date': ytd_high_date.strftime('%Y-%m-%d') if ytd_high_date else None,
            'ytd_low': ytd_low,
            'ytd_low_date': ytd_low_date.strftime('%Y-%m-%d') if ytd_low_date else None,
            'all_time_high': all_time_high,
            'all_time_high_date': all_time_high_date.strftime('%Y-%m-%d') if all_time_high_date else None,
            'all_time_low': all_time_low,
            'all_time_low_date': all_time_low_date.strftime('%Y-%m-%d') if all_time_low_date else None,
        },
        'financial': {
            'revenue': revenue,
            'net_income': net_income,
            'eps': info.get('trailingEps') or info.get('forwardEps'),
            'cash_flow': operating_cashflow,
            'debt_to_equity': info.get('debtToEquity'),
            'roe': roe,
            'gross_margin': gross_margin,
            'net_margin': net_margin,
        },
        'valuation': {
            'pe_ratio': info.get('trailingPE') or info.get('forwardPE'),
            'pb_ratio': info.get('priceToBook'),
            'ps_ratio': info.get('priceToSalesTrailing12Months'),
            'ev_ebitda': info.get('enterpriseToEbitda'),
        },
        'operational': {
            'market_cap': info.get('marketCap'),
            'sector': info.get('sector', 'N/A'),
            'industry': info.get('industry', 'N/A'),
            'employees': info.get('fullTimeEmployees'),
            'business_summary': info.get('longBusinessSummary') or info.get('businessSummary'),
        },
        'macro_sector': {
            'revenue_growth': revenue_growth,
            'earnings_growth': earnings_growth,
            'beta': info.get('beta'),
        },
        'macro_data': data.get('macro_data', {}),
        'earnings_trend': earnings_trend,
        'analyst_data': analyst_data
    }


def export_to_excel(data_dict, output_path):
    """Export data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Analysis"
        
        # Header
        ws['A1'] = f"STOCK FUNDAMENTALS ANALYSIS FOR: {data_dict['ticker']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Analysis Date: {data_dict['analysis_date']}"
        
        row = 4
        
        # Price Information
        ws[f'A{row}'] = "PRICE INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        price_info = data_dict['price_info']
        rsi_value = price_info.get('rsi')
        rsi_signal = price_info.get('rsi_signal', 'N/A')
        rsi_display = f"{rsi_value:.2f} ({rsi_signal})" if rsi_value is not None else 'N/A'
        
        next_earnings = price_info.get('next_earnings_date')
        days_until = price_info.get('days_until_earnings')
        earnings_display = next_earnings if next_earnings else 'N/A'
        if days_until is not None and days_until >= 0:
            earnings_display += f" ({days_until} days)"
        
        price_data = [
            ['Current Price', f"${price_info['current_price']:.2f}" if price_info['current_price'] else 'N/A'],
            ['RSI (14-day)', rsi_display],
            ['Next Earnings Date', earnings_display],
            ['YTD % Change', f"{price_info['ytd_percent_change']:.2f}%" if price_info['ytd_percent_change'] is not None else 'N/A'],
            ['YTD High', f"${price_info['ytd_high']:.2f}" if price_info['ytd_high'] else 'N/A', price_info['ytd_high_date']],
            ['YTD Low', f"${price_info['ytd_low']:.2f}" if price_info['ytd_low'] else 'N/A', price_info['ytd_low_date']],
            ['All-Time High', f"${price_info['all_time_high']:.2f}" if price_info['all_time_high'] else 'N/A', price_info['all_time_high_date']],
            ['All-Time Low', f"${price_info['all_time_low']:.2f}" if price_info['all_time_low'] else 'N/A', price_info['all_time_low_date']],
        ]
        
        for item in price_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if len(item) > 1 else ''
            if len(item) > 2 and item[2]:
                ws[f'C{row}'] = f"({item[2]})"
            row += 1
        
        row += 1
        
        # Financial Metrics
        ws[f'A{row}'] = "FINANCIAL METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        financial = data_dict['financial']
        financial_data = [
            ['Revenue (Sales)', format_number(financial['revenue'])],
            ['Earnings / Net Income', format_number(financial['net_income'])],
            ['Earnings Per Share (EPS)', format_ratio(financial['eps'])],
            ['Cash Flow (Operating)', format_number(financial['cash_flow'])],
            ['Debt-to-Equity Ratio', format_ratio(financial['debt_to_equity'])],
            ['Return on Equity (ROE)', f"{financial['roe']:.2f}%" if financial['roe'] is not None else 'N/A'],
            ['Gross Margin', f"{financial['gross_margin']:.2f}%" if financial['gross_margin'] is not None else 'N/A'],
            ['Net Margin', f"{financial['net_margin']:.2f}%" if financial['net_margin'] is not None else 'N/A'],
        ]
        
        for item in financial_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        row += 1
        
        # Valuation Metrics
        ws[f'A{row}'] = "VALUATION METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        valuation = data_dict['valuation']
        valuation_data = [
            ['P/E (Price-to-Earnings)', format_ratio(valuation['pe_ratio'])],
            ['P/B (Price-to-Book)', format_ratio(valuation['pb_ratio'])],
            ['P/S (Price-to-Sales)', format_ratio(valuation['ps_ratio'])],
            ['EV/EBITDA', format_ratio(valuation['ev_ebitda'])],
        ]
        
        for item in valuation_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        row += 1
        
        # Operational Metrics
        ws[f'A{row}'] = "OPERATIONAL METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        operational = data_dict['operational']
        operational_data = [
            ['Market Capitalization', format_number(operational['market_cap'])],
            ['Sector', operational['sector']],
            ['Industry', operational['industry']],
            ['Full-Time Employees', f"{operational['employees']:,}" if operational['employees'] else 'N/A'],
        ]
        
        for item in operational_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        if operational['business_summary']:
            row += 1
            ws[f'A{row}'] = "Business Summary"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            summary = operational['business_summary'][:500] + "..." if len(operational['business_summary']) > 500 else operational['business_summary']
            ws[f'A{row}'] = summary
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        row += 1
        
        # Macro/Sector Metrics
        ws[f'A{row}'] = "MACROECONOMIC AND SECTOR METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
        row += 1
        
        macro = data_dict['macro_sector']
        fred_data = data_dict.get('macro_data', {})
        
        macro_data = []
        
        # Add FRED data
        if 'interest_rate' in fred_data:
            ir = fred_data['interest_rate']
            macro_data.append([f"Interest Rate ({ir['label']})", f"{ir['value']:.2f}% (as of {ir['date']})"])
        
        if 'inflation' in fred_data:
            inf = fred_data['inflation']
            macro_data.append([f"Inflation ({inf['label']})", f"{inf['value']:.2f}% (as of {inf['date']})"])
        
        if 'unemployment' in fred_data:
            unemp = fred_data['unemployment']
            macro_data.append(['Unemployment Rate', f"{unemp['value']:.2f}% (as of {unemp['date']})"])
        
        if 'gdp' in fred_data:
            gdp = fred_data['gdp']
            macro_data.append(['Real GDP', f"${gdp['value']:.2f}B (as of {gdp['date']})"])
        
        # Add sector data
        macro_data.extend([
            ['Revenue Growth', f"{macro['revenue_growth']:.2f}%" if macro['revenue_growth'] is not None else 'N/A'],
            ['Earnings Growth', f"{macro['earnings_growth']:.2f}%" if macro['earnings_growth'] is not None else 'N/A'],
            ['Beta (Market Correlation)', format_ratio(macro['beta'])],
        ])
        
        for item in macro_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1] if item[1] is not None else 'N/A'
            row += 1
        
        row += 1
        
        # Earnings Trend
        earnings_trend = data_dict.get('earnings_trend', [])
        if earnings_trend:
            ws[f'A{row}'] = "EARNINGS TREND"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
            row += 1
            
            ws[f'A{row}'] = "Period"
            ws[f'B{row}'] = "Earnings"
            ws[f'C{row}'] = "Change"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
            row += 1
            
            for i, period in enumerate(earnings_trend):
                date_str = period['date']
                earnings = period['earnings']
                period_type = period['period']
                
                ws[f'A{row}'] = f"{date_str} ({period_type})"
                ws[f'B{row}'] = format_number(earnings)
                
                # Calculate change
                if i < len(earnings_trend) - 1:
                    prev_earnings = earnings_trend[i + 1]['earnings']
                    if prev_earnings != 0:
                        change = ((earnings - prev_earnings) / abs(prev_earnings)) * 100
                        ws[f'C{row}'] = f"{change:+.2f}%"
                    else:
                        ws[f'C{row}'] = "N/A"
                else:
                    ws[f'C{row}'] = "-"
                
                row += 1
            
            # Overall trend
            if len(earnings_trend) >= 2:
                first_earnings = earnings_trend[-1]['earnings']
                latest_earnings = earnings_trend[0]['earnings']
                if first_earnings != 0:
                    overall_change = ((latest_earnings - first_earnings) / abs(first_earnings)) * 100
                    row += 1
                    ws[f'A{row}'] = "Overall Trend"
                    ws[f'A{row}'].font = Font(bold=True)
                    trend_direction = "↑ Increasing" if overall_change > 0 else "↓ Decreasing" if overall_change < 0 else "→ Stable"
                    ws[f'B{row}'] = f"{trend_direction} ({overall_change:+.2f}%)"
        
        row += 1
        
        # Analyst Ratings
        analyst_data = data_dict.get('analyst_data', {})
        if analyst_data:
            ws[f'A{row}'] = "ANALYST RATINGS & PRICE TARGETS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
            row += 1
            
            if analyst_data.get('average_target'):
                ws[f'A{row}'] = "Average Price Target"
                ws[f'B{row}'] = format_price(analyst_data['average_target'])
                row += 1
                
                if 'upside_downside' in analyst_data:
                    ud = analyst_data['upside_downside']
                    direction = "↑ Upside" if ud > 0 else "↓ Downside" if ud < 0 else "→ Neutral"
                    ws[f'A{row}'] = "Price Target vs Current"
                    ws[f'B{row}'] = f"{direction} ({ud:+.2f}%)"
                    row += 1
                
                if analyst_data.get('high_target'):
                    ws[f'A{row}'] = "High Target"
                    ws[f'B{row}'] = format_price(analyst_data['high_target'])
                    row += 1
                
                if analyst_data.get('low_target'):
                    ws[f'A{row}'] = "Low Target"
                    ws[f'B{row}'] = format_price(analyst_data['low_target'])
                    row += 1
            
            if analyst_data.get('recommendation'):
                rec = analyst_data['recommendation']
                if isinstance(rec, (int, float)):
                    rec_labels = {1: "Strong Buy", 2: "Buy", 3: "Hold", 4: "Underperform", 5: "Sell"}
                    rec_label = rec_labels.get(int(rec), f"{rec:.1f}")
                else:
                    rec_label = str(rec)
                ws[f'A{row}'] = "Average Recommendation"
                ws[f'B{row}'] = rec_label
                row += 1
            
            # Recommendation Breakdown
            if analyst_data.get('recommendation_breakdown'):
                breakdown = analyst_data['recommendation_breakdown']
                for rec_type, count in breakdown.items():
                    if count and count > 0:
                        ws[f'A{row}'] = f"{rec_type}"
                        ws[f'B{row}'] = int(count)
                        row += 1
            
            # Recent Analysts
            if analyst_data.get('analysts'):
                row += 1
                ws[f'A{row}'] = "Recent Analyst Recommendations"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                ws[f'A{row}'] = "Firm"
                ws[f'B{row}'] = "Rating"
                ws[f'C{row}'] = "Date"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].font = Font(bold=True)
                row += 1
                
                for analyst in analyst_data['analysts'][:10]:
                    ws[f'A{row}'] = analyst['firm']
                    ws[f'B{row}'] = analyst['rating']
                    ws[f'C{row}'] = analyst['date']
                    row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        
        wb.save(output_path)
        return True
    except ImportError:
        print(f"{Colors.RED}Error: openpyxl and pandas are required for Excel export. Install with: pip install openpyxl pandas{Colors.END}")
        return False
    except Exception as e:
        print(f"{Colors.RED}Error exporting to Excel: {e}{Colors.END}")
        return False


def export_to_pdf(data_dict, output_path):
    """Export data to PDF file"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0066CC'),
            spaceAfter=30,
        )
        
        title = Paragraph(f"STOCK FUNDAMENTALS ANALYSIS FOR: {data_dict['ticker']}", title_style)
        story.append(title)
        story.append(Paragraph(f"Analysis Date: {data_dict['analysis_date']}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Helper function to create section
        def add_section(title_text, data_list):
            story.append(Paragraph(f"<b>{title_text}</b>", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            table_data = [['Metric', 'Value']]
            for item in data_list:
                if isinstance(item, list) and len(item) >= 2:
                    metric = item[0]
                    value = item[1] if item[1] is not None else 'N/A'
                    if len(item) > 2:
                        value = f"{value} {item[2]}"
                    table_data.append([metric, str(value)])
            
            table = Table(table_data, colWidths=[3*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCFFFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
        
        # Price Information
        price_info = data_dict['price_info']
        rsi_value = price_info.get('rsi')
        rsi_signal = price_info.get('rsi_signal', 'N/A')
        rsi_display = f"{rsi_value:.2f} ({rsi_signal})" if rsi_value is not None else 'N/A'
        
        next_earnings = price_info.get('next_earnings_date')
        days_until = price_info.get('days_until_earnings')
        earnings_display = next_earnings if next_earnings else 'N/A'
        if days_until is not None and days_until >= 0:
            earnings_display += f" ({days_until} days)"
        
        price_data = [
            ['Current Price', f"${price_info['current_price']:.2f}" if price_info['current_price'] else 'N/A'],
            ['RSI (14-day)', rsi_display],
            ['Next Earnings Date', earnings_display],
            ['YTD % Change', f"{price_info['ytd_percent_change']:.2f}%" if price_info['ytd_percent_change'] else 'N/A'],
            ['YTD High', f"${price_info['ytd_high']:.2f}" if price_info['ytd_high'] else 'N/A', f"(on {price_info['ytd_high_date']})" if price_info['ytd_high_date'] else ''],
            ['YTD Low', f"${price_info['ytd_low']:.2f}" if price_info['ytd_low'] else 'N/A', f"(on {price_info['ytd_low_date']})" if price_info['ytd_low_date'] else ''],
            ['All-Time High', f"${price_info['all_time_high']:.2f}" if price_info['all_time_high'] else 'N/A', f"(on {price_info['all_time_high_date']})" if price_info['all_time_high_date'] else ''],
            ['All-Time Low', f"${price_info['all_time_low']:.2f}" if price_info['all_time_low'] else 'N/A', f"(on {price_info['all_time_low_date']})" if price_info['all_time_low_date'] else ''],
        ]
        add_section("PRICE INFORMATION", price_data)
        
        # Financial Metrics
        financial = data_dict['financial']
        financial_data = [
            ['Revenue (Sales)', format_number(financial['revenue'])],
            ['Earnings / Net Income', format_number(financial['net_income'])],
            ['Earnings Per Share (EPS)', format_ratio(financial['eps'])],
            ['Cash Flow (Operating)', format_number(financial['cash_flow'])],
            ['Debt-to-Equity Ratio', format_ratio(financial['debt_to_equity'])],
            ['Return on Equity (ROE)', f"{financial['roe']:.2f}%" if financial['roe'] else 'N/A'],
            ['Gross Margin', f"{financial['gross_margin']:.2f}%" if financial['gross_margin'] else 'N/A'],
            ['Net Margin', f"{financial['net_margin']:.2f}%" if financial['net_margin'] else 'N/A'],
        ]
        add_section("FINANCIAL METRICS", financial_data)
        
        # Valuation Metrics
        valuation = data_dict['valuation']
        valuation_data = [
            ['P/E (Price-to-Earnings)', format_ratio(valuation['pe_ratio'])],
            ['P/B (Price-to-Book)', format_ratio(valuation['pb_ratio'])],
            ['P/S (Price-to-Sales)', format_ratio(valuation['ps_ratio'])],
            ['EV/EBITDA', format_ratio(valuation['ev_ebitda'])],
        ]
        add_section("VALUATION METRICS", valuation_data)
        
        # Operational Metrics
        operational = data_dict['operational']
        operational_data = [
            ['Market Capitalization', format_number(operational['market_cap'])],
            ['Sector', operational['sector']],
            ['Industry', operational['industry']],
            ['Full-Time Employees', f"{operational['employees']:,}" if operational['employees'] else 'N/A'],
        ]
        add_section("OPERATIONAL METRICS", operational_data)
        
        if operational['business_summary']:
            story.append(Paragraph("<b>Business Summary</b>", styles['Heading3']))
            summary = operational['business_summary'][:1000] + "..." if len(operational['business_summary']) > 1000 else operational['business_summary']
            story.append(Paragraph(summary, styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Macro/Sector Metrics
        macro = data_dict['macro_sector']
        fred_data = data_dict.get('macro_data', {})
        
        macro_data = []
        
        # Add FRED data
        if 'interest_rate' in fred_data:
            ir = fred_data['interest_rate']
            macro_data.append([f"Interest Rate ({ir['label']})", f"{ir['value']:.2f}% (as of {ir['date']})"])
        
        if 'inflation' in fred_data:
            inf = fred_data['inflation']
            macro_data.append([f"Inflation ({inf['label']})", f"{inf['value']:.2f}% (as of {inf['date']})"])
        
        if 'unemployment' in fred_data:
            unemp = fred_data['unemployment']
            macro_data.append(['Unemployment Rate', f"{unemp['value']:.2f}% (as of {unemp['date']})"])
        
        if 'gdp' in fred_data:
            gdp = fred_data['gdp']
            macro_data.append(['Real GDP', f"${gdp['value']:.2f}B (as of {gdp['date']})"])
        
        # Add sector data
        macro_data.extend([
            ['Revenue Growth', f"{macro['revenue_growth']:.2f}%" if macro['revenue_growth'] else 'N/A'],
            ['Earnings Growth', f"{macro['earnings_growth']:.2f}%" if macro['earnings_growth'] else 'N/A'],
            ['Beta (Market Correlation)', format_ratio(macro['beta'])],
        ])
        add_section("MACROECONOMIC AND SECTOR METRICS", macro_data)
        
        # Earnings Trend
        earnings_trend = data_dict.get('earnings_trend', [])
        if earnings_trend:
            story.append(Paragraph("<b>EARNINGS TREND</b>", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            trend_table_data = [['Period', 'Earnings', 'Change']]
            for i, period in enumerate(earnings_trend):
                date_str = period['date']
                earnings = period['earnings']
                period_type = period['period']
                
                change_str = ""
                if i < len(earnings_trend) - 1:
                    prev_earnings = earnings_trend[i + 1]['earnings']
                    if prev_earnings != 0:
                        change = ((earnings - prev_earnings) / abs(prev_earnings)) * 100
                        change_str = f"{change:+.2f}%"
                    else:
                        change_str = "N/A"
                else:
                    change_str = "-"
                
                trend_table_data.append([
                    f"{date_str} ({period_type})",
                    format_number(earnings),
                    change_str
                ])
            
            trend_table = Table(trend_table_data, colWidths=[2.5*inch, 2.5*inch, 1*inch])
            trend_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCFFFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(trend_table)
            
            # Overall trend
            if len(earnings_trend) >= 2:
                first_earnings = earnings_trend[-1]['earnings']
                latest_earnings = earnings_trend[0]['earnings']
                if first_earnings != 0:
                    overall_change = ((latest_earnings - first_earnings) / abs(first_earnings)) * 100
                    trend_direction = "↑ Increasing" if overall_change > 0 else "↓ Decreasing" if overall_change < 0 else "→ Stable"
                    story.append(Spacer(1, 0.1*inch))
                    story.append(Paragraph(f"<b>Overall Trend:</b> {trend_direction} ({overall_change:+.2f}%)", styles['Normal']))
            
            story.append(Spacer(1, 0.2*inch))
        
        # Analyst Ratings
        analyst_data = data_dict.get('analyst_data', {})
        if analyst_data:
            story.append(Paragraph("<b>ANALYST RATINGS & PRICE TARGETS</b>", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            analyst_table_data = [['Metric', 'Value']]
            
            if analyst_data.get('average_target'):
                analyst_table_data.append(['Average Price Target', format_price(analyst_data['average_target'])])
                
                if 'upside_downside' in analyst_data:
                    ud = analyst_data['upside_downside']
                    direction = "↑ Upside" if ud > 0 else "↓ Downside" if ud < 0 else "→ Neutral"
                    analyst_table_data.append(['Price Target vs Current', f"{direction} ({ud:+.2f}%)"])
                
                if analyst_data.get('high_target'):
                    analyst_table_data.append(['High Target', format_price(analyst_data['high_target'])])
                
                if analyst_data.get('low_target'):
                    analyst_table_data.append(['Low Target', format_price(analyst_data['low_target'])])
            
            if analyst_data.get('recommendation'):
                rec = analyst_data['recommendation']
                if isinstance(rec, (int, float)):
                    rec_labels = {1: "Strong Buy", 2: "Buy", 3: "Hold", 4: "Underperform", 5: "Sell"}
                    rec_label = rec_labels.get(int(rec), f"{rec:.1f}")
                else:
                    rec_label = str(rec)
                analyst_table_data.append(['Average Recommendation', rec_label])
            
            if analyst_data.get('recommendation_breakdown'):
                breakdown = analyst_data['recommendation_breakdown']
                for rec_type, count in breakdown.items():
                    if count and count > 0:
                        analyst_table_data.append([rec_type, str(int(count))])
            
            analyst_table = Table(analyst_table_data, colWidths=[3*inch, 3*inch])
            analyst_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCFFFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(analyst_table)
            
            # Recent Analysts
            if analyst_data.get('analysts'):
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph("<b>Recent Analyst Recommendations</b>", styles['Heading3']))
                story.append(Spacer(1, 0.1*inch))
                
                analyst_list_data = [['Firm', 'Rating', 'Date']]
                for analyst in analyst_data['analysts'][:10]:
                    analyst_list_data.append([
                        analyst['firm'],
                        analyst['rating'],
                        analyst['date']
                    ])
                
                analyst_list_table = Table(analyst_list_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
                analyst_list_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCFFFF')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(analyst_list_table)
            
            story.append(Spacer(1, 0.2*inch))
        
        doc.build(story)
        return True
    except ImportError:
        print(f"{Colors.RED}Error: reportlab is required for PDF export. Install with: pip install reportlab{Colors.END}")
        return False
    except Exception as e:
        print(f"{Colors.RED}Error exporting to PDF: {e}{Colors.END}")
        return False


